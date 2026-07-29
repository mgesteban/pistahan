#!/usr/bin/env python3
"""Convert the final '2026 Pistahan .xlsx' assignment workbook into the clean
Roster schema for the Google Sheet (see pistahan-checkin-build-brief.md §3).

Supersedes convert_roster.py: the final workbook has a different layout
(AREA / CATEGORY / ASSIGNMENT / DATE / CALL TIME / END TIME / ATTEND
ORIENTATION / VOLUNTEER SOURCE / NAME) and carries NO contact info, so
email/phone/shirt_size are merged in from the earlier roster conversions
by name. Contact sources are read in order; later files win conflicts.

Outputs (all gitignored — they contain PII):
  out/roster_clean.csv    — rows ready to paste into the Roster tab
  out/issues.csv          — rows needing a human decision, with a reason
  out/roster_scanner.json — scanner-app import file (no email/phone)
  out/tokens_issued.json  — persistent name→token map so re-runs NEVER change
                            a token that's already been issued

Usage: python3 scripts/convert_assignments.py "2026 Pistahan .xlsx"
"""
import csv
import json
import re
import secrets
import sys
import unicodedata
from datetime import datetime, time, timezone
from pathlib import Path

import openpyxl

SHEET = 'Volunteer Assignment Alpha Orde'
CONTACT_SOURCES = [
    'backup-2026-07-29/roster_clean.csv',                # full pre-replacement roster
    'backup-2026-07-29/roster_partial_replacement.csv',  # newer partial — wins conflicts
]
# name,size export matching the workbook's names; authoritative for shirt_size
SHIRT_SOURCE = 'tshirt-sizes.csv'
TOKEN_ALPHABET = '23456789ABCDEFGHJKMNPQRSTUVWXYZ'  # no 0/O/1/I/L
DAY_MAP = {
    'SAT': 'SAT',
    'SUN': 'SUN',
    'MON': 'MON',
    'FRI': 'FRI',
    'WEEKEND': 'SAT|SUN',
    'FRI WEEKEND': 'FRI|SAT|SUN',
}
# Spelling variants confirmed by eye against the earlier roster; keys are the
# cleaned lowercase name from the new workbook, values the contact-file name.
ALIASES = {
    'carlito l pantig': 'carlito l. pantig',
    'georgette estilo': 'gigi estilo',
    'joane landayan': 'joanna landayan',
    'ken dare': 'kenneth dare',
    'norman ragasa': 'norman m. ragasa',
    'regina finuliar': 'regina a. finuliar',
}


def norm(name):
    s = unicodedata.normalize('NFKD', name).encode('ascii', 'ignore').decode()
    return re.sub(r'\s+', ' ', s.strip().lower())


def clean_name(raw):
    """Split 'Angela Pedrigal (NO REGISTRATION)' → name + annotation note."""
    annots = re.findall(r'\((.*?)\)', raw) + re.findall(r'"(.*?)"', raw)
    name = re.sub(r'\(.*?\)|".*?"', ' ', raw)
    name = re.sub(r'\s+', ' ', name).strip()
    return name, '; '.join(a.strip() for a in annots if a.strip())


def fmt_time(v):
    if isinstance(v, time):
        return v.strftime('%H:%M')
    return str(v).strip().replace('\n', ' ') if v else ''


def main(path):
    contacts = {}
    for src in CONTACT_SOURCES:
        p = Path(src)
        if not p.exists():
            print(f'WARNING: contact source missing, skipping: {src}')
            continue
        with p.open() as f:
            for r in csv.DictReader(f):
                contacts[norm(f"{r['first_name']} {r['last_name']}")] = r

    shirts = {}
    if Path(SHIRT_SOURCE).exists():
        with open(SHIRT_SOURCE) as f:
            for row in csv.reader(f):
                if not row or not row[0].strip():
                    continue
                name, _ = clean_name(row[0])
                size = row[1].strip().upper() if len(row) > 1 else ''
                if size and size not in ('TBD', 'TSHIRT SIZE'):
                    shirts[norm(name)] = size
    else:
        print(f'WARNING: shirt-size source missing, skipping: {SHIRT_SOURCE}')

    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET]

    out_rows, issues, open_slots = [], [], 0
    seen = {}
    for i, r in enumerate(ws.iter_rows(values_only=True), 1):
        raw_name = str(r[9]).strip() if r[9] else ''
        if not raw_name or raw_name.upper() == 'NAME':
            continue
        if 'open' in raw_name.lower() and len(raw_name) < 12:
            open_slots += 1
            continue

        name, annot = clean_name(raw_name)
        area = str(r[1]).strip().replace('\n', ' ') if r[1] else ''
        category = str(r[2]).strip().replace('\n', ' ') if r[2] else ''
        post = str(r[3]).strip().replace('\n', ' ') if r[3] else ''
        raw_day = str(r[4]).strip().upper().replace('\n', ' ') if r[4] else ''
        days = DAY_MAP.get(raw_day, '')
        shift_start = fmt_time(r[5])
        shift_end = fmt_time(r[6])
        source = str(r[8]).strip().replace('\n', ' ') if r[8] else ''

        team = f'{area} — {category}'.strip(' —')
        assignment = f'{days or raw_day or "?"} {shift_start}–{shift_end}: {team} / {post}'

        # One record per PERSON — one token each; merge extra assignment rows.
        key = norm(name)
        if key in seen:
            rec = seen[key]
            rec['assignments'].append(assignment)
            merged = set(rec['days'].split('|')) | set(days.split('|'))
            rec['days'] = '|'.join(sorted(d for d in merged if d))
            if annot and annot not in rec['notes']:
                rec['notes'] = f"{rec['notes']} / {annot}".strip(' /')
            if not days:
                rec['problems'].append(f'unmapped day: {raw_day!r}')
            continue

        parts = name.split()
        first = ' '.join(parts[:-1]) if len(parts) > 1 else parts[0]
        last = parts[-1] if len(parts) > 1 else ''

        contact = contacts.get(key) or contacts.get(ALIASES.get(key, ''))
        problems = []
        if not last:
            problems.append('single-word name')
        if not days:
            problems.append(f'unmapped day: {raw_day!r}')
        if not contact:
            hint = [
                c for c in contacts
                if last and last.lower() in c.split()
            ]
            problems.append(
                'no contact info (not in earlier roster'
                + (f'; same last name: {hint}' if hint else '') + ')'
            )

        notes = ' / '.join(x for x in (annot, f'src: {source}' if source else '') if x)
        record = {
            'token': (contact or {}).get('token', ''),
            'first_name': first,
            'last_name': last,
            'email': (contact or {}).get('email', ''),
            'phone': (contact or {}).get('phone', ''),
            'shirt_size': shirts.get(key) or (contact or {}).get('shirt_size', ''),
            'team': team,
            'post': post,
            'days': days,
            'shift_start': shift_start,
            'shift_end': shift_end,
            'notes': notes,
            'is_minor': '',
            'source_row': i,
            'assignments': [assignment],
            'problems': problems,
        }
        seen[key] = record
        out_rows.append(record)

    for rec in out_rows:
        rec['assignments'] = ' ; '.join(rec['assignments'])
        probs = rec.pop('problems')
        if rec['email'] == '' and not any('no contact' in p for p in probs):
            probs.append('missing email')
        if probs:
            issues.append({**rec, 'issues': '; '.join(probs)})

    out = Path('out')
    out.mkdir(exist_ok=True)

    # Stable token issuance: existing tokens are never regenerated. Tokens
    # carried over from a contact match are recorded under the new name too.
    cache_path = out / 'tokens_issued.json'
    cache = json.loads(cache_path.read_text()) if cache_path.exists() else {}
    used = set(cache.values())
    for rec in out_rows:
        key = norm(f"{rec['first_name']} {rec['last_name']}")
        if rec['token']:
            cache.setdefault(key, rec['token'])
            used.add(rec['token'])
            continue
        if key not in cache:
            while True:
                t = 'PST-' + ''.join(secrets.choice(TOKEN_ALPHABET) for _ in range(4))
                if t not in used:
                    break
            cache[key] = t
            used.add(t)
        rec['token'] = cache[key]
    cache_path.write_text(json.dumps(cache, indent=1, sort_keys=True))

    # Scanner import file — same shape as the roster endpoint, no email/phone.
    scanner = {
        'version': datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ') + ' (file)',
        'count': len(out_rows),
        'volunteers': [
            {
                'token': r['token'], 'first_name': r['first_name'],
                'last_name': r['last_name'], 'shirt_size': r['shirt_size'],
                'team': r['team'], 'post': r['post'], 'days': r['days'],
                'shift_start': r['shift_start'], 'shift_end': r['shift_end'],
                'notes': r['notes'], 'is_minor': bool(r['is_minor']),
                'assignments': r['assignments'],
            }
            for r in out_rows
        ],
    }
    (out / 'roster_scanner.json').write_text(json.dumps(scanner, indent=1))

    # Column order MUST match ROSTER_COLS in apps-script/Code.gs — the backend
    # reads the Sheet by position.
    sheet_cols = [
        'token', 'first_name', 'last_name', 'email', 'phone', 'shirt_size',
        'team', 'post', 'days', 'shift_start', 'shift_end', 'notes',
        'is_minor', 'assignments', 'pass_sent_email', 'pass_sent_sms',
    ]
    for rec in out_rows:
        rec.setdefault('pass_sent_email', '')
        rec.setdefault('pass_sent_sms', '')

    with open(out / 'roster_clean.csv', 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=sheet_cols, extrasaction='ignore')
        w.writeheader()
        w.writerows(out_rows)
    with open(out / 'issues.csv', 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=sheet_cols + ['source_row', 'issues'])
        w.writeheader()
        w.writerows(issues)

    # One-stop Google Sheet import: all four tabs, roster pre-filled.
    imp = openpyxl.Workbook()
    ws_r = imp.active
    ws_r.title = 'Roster'
    ws_r.append(sheet_cols)
    for rec in out_rows:
        ws_r.append([rec.get(c, '') for c in sheet_cols])
    ws_c = imp.create_sheet('CheckIns')
    ws_c.append(['scan_id', 'timestamp_client', 'timestamp_server', 'token',
                 'station', 'method', 'operator'])
    ws_w = imp.create_sheet('Walkups')
    ws_w.append(['first_name', 'last_name', 'phone', 'shirt_size', 'post',
                 'added_by', 'added_at'])
    ws_d = imp.create_sheet('Dashboard')
    ws_d['A1'] = 'Total checked in'
    ws_d['B1'] = ('=SUMPRODUCT((COUNTIF(CheckIns!D:D,Roster!A2:A400)>0)'
                  '*(Roster!A2:A400<>""))')
    ws_d['A2'] = 'Roster total'
    ws_d['B2'] = '=COUNTA(Roster!A2:A400)'
    ws_d['A3'] = 'Percent'
    ws_d['B3'] = '=IF(B2=0,0,B1/B2)'
    ws_d['A5'] = 'STILL MISSING (name, post):'
    ws_d['A6'] = ('=FILTER(Roster!B2:C400,'
                  'COUNTIF(CheckIns!D:D,Roster!A2:A400)=0,Roster!A2:A400<>"")')
    imp.save(out / 'pistahan_sheet_import.xlsx')

    with_contact = sum(1 for r in out_rows if r['email'] or r['phone'])
    print(f'{len(out_rows)} volunteers -> out/roster_clean.csv')
    print(f'  {with_contact} with contact info, {len(out_rows) - with_contact} without')
    print(f'  {open_slots} "(Open)" slots skipped')
    print(f'{len(issues)} rows need review -> out/issues.csv')
    print('Google Sheet import file -> out/pistahan_sheet_import.xlsx')


if __name__ == '__main__':
    main(sys.argv[1] if len(sys.argv) > 1 else '2026 Pistahan .xlsx')
