#!/usr/bin/env python3
"""Convert the working 'Pistahan 33 Volunteer Roster.xlsx' into the clean
Roster schema for the Google Sheet (see pistahan-checkin-build-brief.md §3).

Outputs (all gitignored — they contain PII):
  out/roster_clean.csv    — rows ready to paste into the Roster tab (tokens filled)
  out/issues.csv          — rows needing a human decision, with a reason
  out/roster_scanner.json — scanner-app import file (no email/phone)
  out/tokens_issued.json  — persistent name→token map so re-runs NEVER change
                            a token that's already been issued. Once the roster
                            is imported to the Google Sheet, the Sheet's token
                            column is authoritative; this cache mirrors it.

Usage: python3 scripts/convert_roster.py "Pistahan 33 Volunteer Roster.xlsx"
"""
import csv
import json
import re
import secrets
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

TOKEN_ALPHABET = '23456789ABCDEFGHJKMNPQRSTUVWXYZ'  # no 0/O/1/I/L

SHIRT_COLS = {13: 'S', 14: 'M', 15: 'L', 16: 'XL', 17: '2XL', 18: '3XL'}
DAY_MAP = {
    'SAT': 'SAT',
    'SUN': 'SUN',
    'WEEKEND': 'SAT|SUN',
    'FRI': 'FRI',
    'FRI\nWEEKEND': 'FRI|SAT|SUN',
}


def norm_phone(raw):
    if raw is None:
        return '', 'missing phone'
    s = str(raw).strip()
    if s.endswith('.0'):  # Excel stored it as a float
        s = s[:-2]
    digits = re.sub(r'\D', '', s)
    if len(digits) == 11 and digits.startswith('1'):
        digits = digits[1:]
    if len(digits) == 10:
        return f'+1{digits}', None
    return s, f'bad phone: {raw!r}'


def norm_email(raw):
    if raw is None:
        return '', 'missing email'
    s = str(raw).strip().lower()
    if '@' not in s or '.' not in s.split('@')[-1]:
        return s, f'bad email: {raw!r}'
    if s.endswith('.con') or s.endswith('.comm') or s.endswith('.cmo'):
        return s, f'suspicious email tld: {raw!r}'
    return s, None


def main(path):
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    out_rows, issues = [], []
    seen_names = {}
    area = category = ''

    for i, r in enumerate(rows[1:], start=2):
        # Section headers carry AREA/CATEGORY; summary rows carry junk like
        # "SUBTOTAL" — never let those bleed into volunteer rows below them.
        for col, target in ((0, 'area'), (1, 'category')):
            v = str(r[col]).strip().replace('\n', ' ') if r[col] else ''
            if v and 'TOTAL' not in v.upper():
                if target == 'area':
                    area = v
                else:
                    category = v

        name = str(r[8]).strip() if r[8] else ''
        if not name or name.upper() == 'NAME':
            continue

        problems = []
        parts = name.split()
        first = ' '.join(parts[:-1]) if len(parts) > 1 else parts[0]
        last = parts[-1] if len(parts) > 1 else ''
        if not last:
            problems.append('single-word name')

        email, e_issue = norm_email(r[9])
        if e_issue:
            problems.append(e_issue)
        phone, p_issue = norm_phone(r[10])
        if p_issue:
            problems.append(p_issue)

        marks = [label for col, label in SHIRT_COLS.items() if r[col]]
        shirt = marks[0] if len(marks) == 1 else ''
        if not marks:
            problems.append('no shirt size')
        elif len(marks) > 1:
            problems.append(f'multiple shirt sizes: {marks}')

        raw_day = str(r[3]).strip().upper() if r[3] else ''
        days = DAY_MAP.get(raw_day.replace('\n', '\n'), '')
        if not days:
            problems.append(f'unmapped day: {raw_day!r}')

        notes = ' / '.join(
            str(x).strip().replace('\n', ' ') for x in (r[11], r[12]) if x
        )
        post = str(r[2]).strip().replace('\n', ' ') if r[2] else ''
        shift_start = str(r[4]).strip().replace('\n', ' ') if r[4] else ''
        shift_end = str(r[5]).strip().replace('\n', ' ') if r[5] else ''

        team = f'{area} — {category}'.strip(' —')
        assignment = f'{days or "?"} {shift_start}–{shift_end}: {team} / {post}'

        # One record per PERSON — 55 people hold multiple assignments, and a
        # person gets exactly one token. Merge subsequent rows into the first.
        key = re.sub(r'\s+', ' ', name.lower())
        if key in seen_names:
            rec = seen_names[key]
            rec['assignments'].append(assignment)
            merged = set(rec['days'].split('|')) | set(days.split('|'))
            rec['days'] = '|'.join(sorted(d for d in merged if d))
            for field, value, problem in (
                ('email', email, e_issue),
                ('phone', phone, p_issue),
                ('shirt_size', shirt, None),
            ):
                if not rec[field] and value and not problem:
                    rec[field] = value
            if notes and notes not in rec['notes']:
                rec['notes'] = f"{rec['notes']} / {notes}".strip(' /')
            continue

        record = {
            'token': '',  # filled from the persistent token cache below
            'first_name': first,
            'last_name': last,
            'email': email,
            'phone': phone,
            'shirt_size': shirt,
            'team': team,
            'post': post,
            'days': days,
            'shift_start': shift_start,
            'shift_end': shift_end,
            'notes': notes,
            'is_minor': '',  # not captured in source sheet; fill manually
            'source_row': i,
            'assignments': [assignment],
            'problems': problems,
        }
        seen_names[key] = record
        out_rows.append(record)

    # Finalize per-person records: merges above may have filled gaps, so
    # compute completeness issues only now, from the final state.
    for rec in out_rows:
        rec['assignments'] = ' ; '.join(rec['assignments'])
        probs = [
            p for p in rec.pop('problems')
            if not p.startswith(('missing', 'no shirt', 'unmapped day'))
        ]
        if not rec['email']:
            probs.append('missing email')
        if not rec['phone']:
            probs.append('missing phone')
        if not rec['shirt_size']:
            probs.append('no shirt size')
        if not rec['days']:
            probs.append('no day mapped')
        if probs:
            issues.append({**rec, 'issues': '; '.join(probs)})

    out = Path('out')
    out.mkdir(exist_ok=True)

    # Stable token issuance: existing tokens are never regenerated.
    cache_path = out / 'tokens_issued.json'
    cache = json.loads(cache_path.read_text()) if cache_path.exists() else {}
    used = set(cache.values())
    for rec in out_rows:
        key = f"{rec['first_name']} {rec['last_name']}".lower()
        if key not in cache:
            while True:
                t = 'PST-' + ''.join(secrets.choice(TOKEN_ALPHABET) for _ in range(4))
                if t not in used:
                    break
            cache[key] = t
            used.add(t)
        rec['token'] = cache[key]
    cache_path.write_text(json.dumps(cache, indent=1, sort_keys=True))

    # Scanner import file — same shape as the roster endpoint, no email/phone.
    scanner = {
        'version': datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ') + ' (file)',
        'count': len(out_rows),
        'volunteers': [
            {
                'token': r['token'], 'first_name': r['first_name'],
                'last_name': r['last_name'], 'shirt_size': r['shirt_size'],
                'team': r['team'], 'post': r['post'], 'days': r['days'],
                'shift_start': r['shift_start'], 'shift_end': r['shift_end'],
                'notes': r['notes'], 'is_minor': bool(r['is_minor']),
                'assignments': r['assignments'],
            }
            for r in out_rows
        ],
    }
    (out / 'roster_scanner.json').write_text(json.dumps(scanner, indent=1))

    # Column order MUST match ROSTER_COLS in apps-script/Code.gs — the backend
    # reads the Sheet by position.
    sheet_cols = [
        'token', 'first_name', 'last_name', 'email', 'phone', 'shirt_size',
        'team', 'post', 'days', 'shift_start', 'shift_end', 'notes',
        'is_minor', 'assignments', 'pass_sent_email', 'pass_sent_sms',
    ]
    for rec in out_rows:
        rec.setdefault('pass_sent_email', '')
        rec.setdefault('pass_sent_sms', '')

    with open(out / 'roster_clean.csv', 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=sheet_cols, extrasaction='ignore')
        w.writeheader()
        w.writerows(out_rows)
    with open(out / 'issues.csv', 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=sheet_cols + ['source_row', 'issues'])
        w.writeheader()
        w.writerows(issues)

    # One-stop Google Sheet import: all four tabs, roster pre-filled.
    imp = openpyxl.Workbook()
    ws_r = imp.active
    ws_r.title = 'Roster'
    ws_r.append(sheet_cols)
    for rec in out_rows:
        ws_r.append([rec.get(c, '') for c in sheet_cols])
    ws_c = imp.create_sheet('CheckIns')
    ws_c.append(['scan_id', 'timestamp_client', 'timestamp_server', 'token',
                 'station', 'method', 'operator'])
    ws_w = imp.create_sheet('Walkups')
    ws_w.append(['first_name', 'last_name', 'phone', 'shirt_size', 'post',
                 'added_by', 'added_at'])
    ws_d = imp.create_sheet('Dashboard')
    ws_d['A1'] = 'Total checked in'
    ws_d['B1'] = ('=SUMPRODUCT((COUNTIF(CheckIns!D:D,Roster!A2:A400)>0)'
                  '*(Roster!A2:A400<>""))')
    ws_d['A2'] = 'Roster total'
    ws_d['B2'] = '=COUNTA(Roster!A2:A400)'
    ws_d['A3'] = 'Percent'
    ws_d['B3'] = '=IF(B2=0,0,B1/B2)'
    ws_d['A5'] = 'STILL MISSING (name, post):'
    ws_d['A6'] = ('=FILTER(Roster!B2:C400,'
                  'COUNTIF(CheckIns!D:D,Roster!A2:A400)=0,Roster!A2:A400<>"")')
    imp.save(out / 'pistahan_sheet_import.xlsx')

    print(f'{len(out_rows)} volunteers -> out/roster_clean.csv')
    print(f'{len(issues)} rows need review -> out/issues.csv')
    print('Google Sheet import file -> out/pistahan_sheet_import.xlsx')


if __name__ == '__main__':
    main(sys.argv[1] if len(sys.argv) > 1 else 'Pistahan 33 Volunteer Roster.xlsx')
